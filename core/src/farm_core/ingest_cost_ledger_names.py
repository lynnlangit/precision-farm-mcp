"""Extract just the Field column from every season tab of the farmer's cost
ledger workbook -- tolerant of the specific messes the ledger is known to
contain (a merged header row, a totals row inside the data range, blank rows
mid-table). This is a narrow, Phase-1-only reader (field names + acres, for
identity resolution); ingest_cost_ledger.py does the full-column read with a
confirmed mapping.
"""

from __future__ import annotations

from pathlib import Path

import duckdb
import openpyxl

from . import db as db_mod
from . import xlsx_rows


def _extract_field_names(rows: list[tuple]) -> list[tuple[int, str, float]]:
    header_idx = xlsx_rows.find_header_row(rows)
    out: list[tuple[int, str, float]] = []
    for row_index, row in xlsx_rows.iter_data_rows(rows, header_idx):
        acres = row[2]  # column C: Field, Crop, Acres, ...
        out.append((row_index, str(row[0]), float(acres)))
    return out


def ingest_cost_ledger_field_names(con: duckdb.DuckDBPyConnection, data_dir: Path) -> int:
    """Ingest the Field column of every season tab in cost_ledger.xlsx.
    Returns the number of rows ingested.
    """
    path = data_dir / "cost_ledger.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    total = 0
    for sheet_name in wb.sheetnames:
        if not sheet_name.isdigit():
            continue  # skip the "Unit Prices" tab
        season = int(sheet_name)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        names = _extract_field_names(rows)

        db_rows = [
            {
                "season": season,
                "raw_field_name": name,
                "acres": acres,
                "row_index": row_index,
                "source_file": f"cost_ledger.xlsx::{sheet_name}",
            }
            for row_index, name, acres in names
        ]
        db_mod.replace_rows(
            con, "cost_ledger_field_names", f"cost_ledger.xlsx::{sheet_name}", db_rows
        )
        total += len(db_rows)

    return total
