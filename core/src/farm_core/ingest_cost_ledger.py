"""Full-column ingestion of the cost ledger workbook. Every farmer's ledger is
different, so nothing here assumes fixed column positions -- it proposes a
mapping from the header row's text, and that mapping goes through
confirm_fn before a single data row is read. A mapping is keyed by the exact
header text it was derived from, so every season sharing the same header
shape reuses the one confirmed mapping instead of asking again; a season with
different headers (a different cost basis, in this data) gets its own
proposal and its own confirmation.

Costs are normalized to $/ac at ingestion time so every downstream reader
sees a consistent unit -- the original declared basis is preserved in the
cost_basis column for provenance, not lost.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import duckdb
import openpyxl

from . import confirm as confirm_mod
from . import db as db_mod
from . import xlsx_rows

_COLUMN_LABELS = {
    "field_name_col": "Field",
    "crop_col": "Crop",
    "acres_col": "Acres",
    "seed_col": "Seed",
    "fertilizer_col": "Fertilizer",
    "chemical_col": "Chemical",
    "fuel_col": "Fuel",
    "cash_rent_col": "Cash Rent",
    "notes_col": "Notes",
}

_COST_COLUMNS = ("seed_col", "fertilizer_col", "chemical_col", "fuel_col", "cash_rent_col")


def _parse_basis(header_cell: str) -> str | None:
    if "(Total $)" in header_cell:
        return "total_dollars"
    if "($/ac)" in header_cell:
        return "per_acre"
    return None


def propose_column_mapping(header_row: tuple) -> tuple[dict[str, Any], str]:
    """Return (proposal, confidence) for a header row. The proposal maps each
    canonical field to a column index plus the declared cost basis.
    """
    header_texts = [str(c) if c is not None else "" for c in header_row]
    mapping: dict[str, Any] = {}
    bases_seen: set[str] = set()

    for canonical, label in _COLUMN_LABELS.items():
        matches = [i for i, text in enumerate(header_texts) if text.startswith(label)]
        if len(matches) != 1:
            return {
                "error": f"expected exactly one column starting with {label!r}, "
                f"found {len(matches)}"
            }, "low"
        col_index = matches[0]
        mapping[canonical] = col_index
        if canonical in _COST_COLUMNS:
            basis = _parse_basis(header_texts[col_index])
            if basis is None:
                error = f"could not determine cost basis from {header_texts[col_index]!r}"
                return {"error": error}, "low"
            bases_seen.add(basis)

    if len(bases_seen) != 1:
        return {"error": f"cost columns disagree on basis: {bases_seen}"}, "low"

    mapping["cost_basis"] = next(iter(bases_seen))
    return mapping, "high"


def _mapping_key(header_row: tuple) -> str:
    normalized = tuple(str(c) if c is not None else "" for c in header_row)
    return "column_mapping:" + "|".join(normalized)


def _confirmed_mapping_for_season(
    rows: list[tuple], season: int, confirm_fn: confirm_mod.ConfirmFn
) -> tuple[dict[str, Any], int]:
    header_idx = xlsx_rows.find_header_row(rows)
    header_row = rows[header_idx]
    proposal, confidence = propose_column_mapping(header_row)

    if "error" in proposal:
        raise confirm_mod.ConfirmationRejected(
            f"Cost ledger column mapping for season {season} is ambiguous: {proposal['error']}"
        )

    request = confirm_mod.ConfirmationRequest(
        kind="column_mapping",
        key=_mapping_key(header_row),
        subject=f"Cost ledger column mapping for season {season} (header: {header_row})",
        proposal=proposal,
        confidence=confidence,
        context={"header_row": list(header_row), "season": season},
    )
    response = confirm_fn(request)
    if not response.approved:
        raise confirm_mod.ConfirmationRejected(
            f"Cost ledger column mapping for season {season} was not confirmed"
        )
    return response.answer, header_idx


def _extract_rows(rows: list[tuple], header_idx: int, mapping: dict[str, Any]) -> list[dict]:
    basis = mapping["cost_basis"]
    out = []
    for row_index, row in xlsx_rows.iter_data_rows(rows, header_idx):
        acres = float(row[mapping["acres_col"]])
        raw_costs = {c: float(row[mapping[c]]) for c in _COST_COLUMNS}
        if basis == "total_dollars":
            raw_costs = {c: v / acres for c, v in raw_costs.items()}

        out.append(
            {
                "row_index": row_index,
                "raw_field_name": str(row[mapping["field_name_col"]]),
                "crop": str(row[mapping["crop_col"]]),
                "acres": acres,
                "seed_cost_per_ac": raw_costs["seed_col"],
                "fertilizer_cost_per_ac": raw_costs["fertilizer_col"],
                "chemical_cost_per_ac": raw_costs["chemical_col"],
                "fuel_cost_per_ac": raw_costs["fuel_col"],
                "cash_rent_per_ac": raw_costs["cash_rent_col"],
                "cost_basis": basis,
                "notes": row[mapping["notes_col"]],
            }
        )
    return out


def ingest_cost_ledger(
    con: duckdb.DuckDBPyConnection, data_dir: Path, confirm_fn: confirm_mod.ConfirmFn
) -> int:
    """Full-column ingest of every season tab in cost_ledger.xlsx. Returns the
    number of rows ingested.
    """
    path = data_dir / "cost_ledger.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    total = 0
    for sheet_name in wb.sheetnames:
        if not sheet_name.isdigit():
            continue
        season = int(sheet_name)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        mapping, header_idx = _confirmed_mapping_for_season(rows, season, confirm_fn)
        mapping_version = _mapping_key(rows[header_idx])
        extracted = _extract_rows(rows, header_idx, mapping)

        source_file = f"cost_ledger.xlsx::{sheet_name}"
        db_rows = [
            {
                **row,
                "season": season,
                "mapping_version": mapping_version,
                "source_file": source_file,
            }
            for row in extracted
        ]
        db_mod.replace_rows(con, "cost_ledger_rows", f"cost_ledger.xlsx::{sheet_name}", db_rows)
        total += len(db_rows)

    return total
